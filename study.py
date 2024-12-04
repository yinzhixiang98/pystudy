from flask import Flask, request, render_template_string
import requests
import os
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    try:
        if request.method == 'POST':
            excel_file = request.files['excel_file']
            if excel_file:
                excel_path = os.path.join('D:/pyworkspace/study/execl/', excel_file.filename)
                excel_file.save(excel_path)

                workbook = load_workbook(excel_path)
                worksheet = workbook['SQL Results']
                rows_list = []

                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    rows_list.append(list(row))

                download_dir = 'D:/Downloads'
                if not os.path.exists(download_dir):
                    os.makedirs(download_dir)

                for row in rows_list:
                    no = row[0]
                    url = row[3]
                    fileName = row[4]
                    fileType = row[5]
                    idN = row[6]
                    print(idN, fileName)
                    concatenated_string = "_".join([idN, fileName])
                    result = ".".join([concatenated_string, fileType])
                    local_path = os.path.join(download_dir, result)

                    response = requests.get(url, stream=True)

                    if response.status_code == 200:
                        with open(local_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                f.write(chunk)
                        print(f'文件 {url} 已下载并重命名为：{result}')
                    else:
                        print(f'文件 {url} 下载失败，状态码：{response.status_code}')

                return '文件已上传并处理完成。'
    except Exception as e:
        print(f"发生错误：{e}")
        return f"发生错误：{e}"

    return '''
    <!doctype html>
    <title>文件上传</title>
    <h1>上传Excel文件</h1>
    <form method=post enctype=multipart/form-data>
      <input type=file name=excel_file>
      <input type=submit value=上传>
    </form>
    '''


if __name__ == '__main__':
    app.run(debug=True)